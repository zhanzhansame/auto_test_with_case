import unittest

from services.export_service import build_testcases_csv, build_testcases_markdown, build_testcases_xlsx_bytes


class TestExportService(unittest.TestCase):
    def test_testcases_xlsx_step_expected_columns_and_alignment(self):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl 未安装，跳过 xlsx 导出测试")

        modules = [
            {
                "level1_module": "L1",
                "level2_module": "L2",
                "level3_function": "F",
                "testCases": [
                    {
                        "case_id": "C1",
                        "case_name": "验证1",
                        "preconditions": "P1",
                        "summary": "S1",
                        "steps": ["s1", "s2"],
                        "expected_results": ["e1", "e2"],
                    },
                    {
                        "case_id": "C2",
                        "case_name": "验证2",
                        "preconditions": "P2",
                        "summary": "S2",
                        "steps": ["s1_only"],
                        "expected_results": [],
                    },
                ],
            }
        ]

        buffer = build_testcases_xlsx_bytes(modules)
        wb = load_workbook(filename=buffer)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        self.assertIn("step", headers)
        self.assertIn("expected_result", headers)

        step_col = headers.index("step") + 1
        exp_col = headers.index("expected_result") + 1

        self.assertEqual(ws.cell(row=2, column=step_col).value, "s1")
        self.assertEqual(ws.cell(row=2, column=exp_col).value, "e1")
        self.assertEqual(ws.cell(row=3, column=step_col).value, "s2")
        self.assertEqual(ws.cell(row=3, column=exp_col).value, "e2")
        self.assertEqual(ws.cell(row=4, column=step_col).value, "s1_only")
        self.assertEqual(ws.cell(row=4, column=exp_col).value, "")

        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        expected_ranges = {"A2:A3", "B2:B3", "C2:C3", "D2:D3", "E2:E3", "F2:F3"}
        self.assertTrue(expected_ranges.issubset(merged_ranges))
        self.assertIsNone(ws.cell(row=3, column=1).value)

    def test_testcases_csv_step_expected_rows(self):
        modules = [
            {
                "level1_module": "L1",
                "level2_module": "L2",
                "level3_function": "F",
                "testCases": [
                    {
                        "case_id": "C1",
                        "case_name": "验证1",
                        "preconditions": "P1",
                        "summary": "S1",
                        "steps": ["s1", "s2"],
                        "expected_results": ["e1", "e2"],
                    }
                ],
            }
        ]

        csv_text = build_testcases_csv(modules)
        self.assertIn("step,expected_result", csv_text)
        self.assertIn("L2,F,C1,验证1,P1,S1,s1,e1", csv_text)
        self.assertIn("L2,F,C1,验证1,P1,S1,s2,e2", csv_text)

    def test_testcases_markdown_contains_step_expected_pairs(self):
        modules = [
            {
                "level1_module": "L1",
                "level2_module": "L2",
                "level3_function": "F",
                "testCases": [
                    {
                        "case_id": "C1",
                        "case_name": "验证1",
                        "preconditions": "",
                        "summary": "",
                        "steps": ["s1", "s2"],
                        "expected_results": ["e1", "e2"],
                    }
                ],
            }
        ]

        md = build_testcases_markdown(modules)
        self.assertIn("| s1 | e1 |", md)
        self.assertIn("| s2 | e2 |", md)


if __name__ == "__main__":
    unittest.main()

