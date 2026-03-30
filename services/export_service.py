import csv
import io
from typing import Any, Dict, List, Tuple



def _build_csv_string(header: List[str], rows: List[List[Any]]) -> str:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(header)
    writer.writerows(rows)
    return output.getvalue()


def _build_xlsx_bytes(
    header: List[str],
    rows: List[List[Any]],
    sheet_name: str = "Sheet1",
    merge_ranges: List[Tuple[int, int, int, int]] = None,
) -> io.BytesIO:
    # openpyxl 可能在某些轻量环境里没安装；仅当真正导出 xlsx 时才需要它
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "openpyxl 未安装，无法导出 xlsx。请安装 autotest_py/requirements.txt"
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(header)
    for row in rows:
        worksheet.append(["" if value is None else value for value in row])

    if merge_ranges:
        for start_col, start_row, end_col, end_row in merge_ranges:
            worksheet.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col,
            )

    # 让步骤/预期这类长文本在单元格内自动换行显示
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for r in range(2, worksheet.max_row + 1):
        for c in range(1, worksheet.max_column + 1):
            worksheet.cell(row=r, column=c).alignment = wrap_align

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_requirements_markdown(modules: List[Dict[str, Any]]) -> str:
    lines: List[str] = []
    for module in modules:
        level1 = str(module.get("level1_module", "未分类")).strip() or "未分类"
        level2 = str(module.get("level2_module", "未分类")).strip() or "未分类"
        level3 = str(module.get("level3_function", "未分类")).strip() or "未分类"
        content = str(module.get("content", "")).strip()
        lines.extend([f"# {level1}", f"## {level2}", f"### {level3}", content, ""])
    return "\n".join(lines)


def build_requirements_csv(modules: List[Dict[str, Any]]) -> str:
    header = ["id", "level1_module", "level2_module", "level3_function", "content"]
    rows: List[List[Any]] = []
    for module in modules:
        rows.append(
            [
                module.get("id", ""),
                module.get("level1_module", ""),
                module.get("level2_module", ""),
                module.get("level3_function", ""),
                module.get("content", ""),
            ]
        )
    return _build_csv_string(header, rows)


def build_requirements_xlsx_bytes(modules: List[Dict[str, Any]]) -> io.BytesIO:
    header = ["id", "level1_module", "level2_module", "level3_function", "content"]
    rows: List[List[Any]] = []
    for module in modules:
        rows.append(
            [
                module.get("id", ""),
                module.get("level1_module", ""),
                module.get("level2_module", ""),
                module.get("level3_function", ""),
                module.get("content", ""),
            ]
        )
    return _build_xlsx_bytes(header, rows, sheet_name="Requirements")


def build_testcases_csv(modules: List[Dict[str, Any]]) -> str:
    """
    导出 CSV 时按“每个步骤/预期结果一行”输出，保持步骤和预期结果按顺序一一对应。
    """
    header = [
        "module",
        "function",
        "case_id",
        "case_name",
        "preconditions",
        "summary",
        "step",
        "expected_result",
    ]
    rows: List[List[Any]] = []
    for module in modules:
        module_name = module.get("level2_module") or module.get("level1_module") or ""
        function_name = module.get("level3_function") or module.get("level2_module") or ""

        test_cases = module.get("testCases")
        if isinstance(test_cases, list) and test_cases:
            for tc in test_cases:
                steps = tc.get("steps", []) or []
                expected_results = tc.get("expected_results", []) or []
                max_len = max(1, len(steps), len(expected_results))
                for i in range(max_len):
                    rows.append(
                        [
                            module_name,
                            function_name,
                            tc.get("case_id", ""),
                            tc.get("case_name", ""),
                            tc.get("preconditions", ""),
                            tc.get("summary", ""),
                            steps[i] if i < len(steps) else "",
                            expected_results[i] if i < len(expected_results) else "",
                        ]
                    )
        else:
            test_points = module.get("testPoints")
            if isinstance(test_points, list) and test_points:
                for idx, point in enumerate(test_points, start=1):
                    rows.append([module_name, function_name, f"C{idx}", str(point), "", "", "", ""])

    return _build_csv_string(header, rows)


def _normalize_list(value: Any) -> List[str]:
    if isinstance(value, list):
        return [str(v) for v in value]
    return []


def _iter_testcases(modules: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    将 modules 统一铺平成“可导出的用例记录”，便于后续导出 csv/xlsx/md。
    """
    records: List[Dict[str, Any]] = []
    for module in modules:
        module_name = module.get("level2_module") or module.get("level1_module") or ""
        function_name = module.get("level3_function") or module.get("level2_module") or ""

        test_cases = module.get("testCases")
        if isinstance(test_cases, list) and test_cases:
            for tc in test_cases:
                steps = _normalize_list(tc.get("steps"))
                expected_results = _normalize_list(tc.get("expected_results"))
                records.append(
                    {
                        "module": module_name,
                        "function": function_name,
                        "case_id": tc.get("case_id", ""),
                        "case_name": tc.get("case_name", ""),
                        "preconditions": tc.get("preconditions", ""),
                        "summary": tc.get("summary", ""),
                        "steps": steps,
                        "expected_results": expected_results,
                    }
                )
        else:
            test_points = module.get("testPoints")
            if isinstance(test_points, list) and test_points:
                for idx, point in enumerate(test_points, start=1):
                    records.append(
                        {
                            "module": module_name,
                            "function": function_name,
                            "case_id": f"C{idx}",
                            "case_name": str(point),
                            "preconditions": "",
                            "summary": "",
                            "steps": [],
                            "expected_results": [],
                        }
                    )
    return records


def build_testcases_markdown(modules: List[Dict[str, Any]]) -> str:
    """
    按“模块/功能 -> 每条用例 -> 步骤-预期一一对应表格”输出。
    """
    lines: List[str] = []
    records = _iter_testcases(modules)

    # 用例可能来自不同 module/function，这里按 module/function 分组渲染更好读
    groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for r in records:
        key = (r.get("module", ""), r.get("function", ""))
        groups.setdefault(key, []).append(r)

    for (module_name, function_name), group_records in groups.items():
        if module_name:
            lines.append(f"# {module_name}")
        if function_name:
            lines.append(f"## {function_name}")
        lines.append("")

        for idx, tc in enumerate(group_records, start=1):
            case_id = tc.get("case_id", "") or f"C{idx}"
            case_name = tc.get("case_name", "") or ""
            lines.append(f"### 用例 {case_id} {case_name}".strip())
            preconditions = str(tc.get("preconditions", "") or "").strip()
            summary = str(tc.get("summary", "") or "").strip()
            if preconditions:
                lines.append(f"- 前置条件：{preconditions}")
            if summary:
                lines.append(f"- 摘要：{summary}")
            lines.append("")

            steps = tc.get("steps", []) or []
            expected_results = tc.get("expected_results", []) or []
            max_len = max(0, max(len(steps), len(expected_results)))

            # 保证“步骤和预期结果按序一一对应”
            lines.append("| 步骤 | 预期结果 |")
            lines.append("|---|---|")
            for i in range(max_len):
                step_val = steps[i] if i < len(steps) else ""
                exp_val = expected_results[i] if i < len(expected_results) else ""
                lines.append(f"| {step_val} | {exp_val} |")

            lines.append("")

    return "\n".join(lines).strip() + "\n"


def build_testcases_xlsx_bytes(modules: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Excel 导出按“每一步骤一行”输出，并保证步骤与预期结果一一对应。
    对同一用例的前置信息字段做向下合并，避免重复显示。
    """
    records = _iter_testcases(modules)
    header = [
        "module",
        "function",
        "case_id",
        "case_name",
        "preconditions",
        "summary",
        "step",
        "expected_result",
    ]

    rows: List[List[Any]] = []
    merge_ranges: List[Tuple[int, int, int, int]] = []
    current_row = 2

    for tc in records:
        steps = tc.get("steps", []) or []
        expected_results = tc.get("expected_results", []) or []
        max_len = max(1, len(steps), len(expected_results))
        start_row = current_row

        for i in range(max_len):
            rows.append(
                [
                    tc.get("module", "") if i == 0 else "",
                    tc.get("function", "") if i == 0 else "",
                    tc.get("case_id", "") if i == 0 else "",
                    tc.get("case_name", "") if i == 0 else "",
                    tc.get("preconditions", "") if i == 0 else "",
                    tc.get("summary", "") if i == 0 else "",
                    steps[i] if i < len(steps) else "",
                    expected_results[i] if i < len(expected_results) else "",
                ]
            )
            current_row += 1

        if max_len > 1:
            for col in range(1, 7):
                merge_ranges.append((col, start_row, col, current_row - 1))

    return _build_xlsx_bytes(
        header,
        rows,
        sheet_name="TestCases",
        merge_ranges=merge_ranges,
    )

